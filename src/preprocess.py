import os
import time

import jieba
import openpyxl
from gensim.models import Word2Vec

from path import *


def read_excel(filename, sheet_name):
    """
    read sheet (default as Sheet1) of the filename from global input path "./data".
    :return [{label : value}], key as the labels, for every row.
    """
    os.chdir(root_path + data_path)

    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.get_sheet_by_name(sheet_name)

    num_row = sheet.max_row
    num_col = sheet.max_column

    return_list = []
    labels = sheet[1]
    for row in range(2, num_row):
        curr_row = sheet[row]
        row_vector = {}
        for col in range(num_col):
            row_vector[labels[col].value] = curr_row[col].value
        return_list.append(row_vector)

    print("Excel file read finished.")
    return return_list


def extract_label(data, key):
    """
    Extract the values of colum key in data.
    :param data: data representative a sheet, returned by read_excel.
    :param key: the column label.
    :return: a list of values in the column of the label.
    """
    sentences = []
    for vector in data:
        sentences.append(vector[key])
    print("Column extracted with key \"" + key + "\"")
    return sentences


"""
Preload the stopwords to avoid to much IO operation.
"""
stopwords = set(open(root_path + data_path + "stopwords.txt", mode='r').read().split())


def segment_and_clear(corpus):
    """
    Do the segmentation and remove stop words, according to HIT Stop Word List.
    :param corpus: the list of raw texts, extracted from the sheet.
    :return: list that has lists of words processed.
    """
    return_list = []
    for text in corpus:
        return_list.append(list(filter(lambda w: False if w in stopwords else True, jieba.lcut(text))))
    print("Segmentation and remove stop words finished.")
    return return_list


def train_word2vec_model(corpus_words, vector_size=100, window=5, min_count=1, workers=4):
    """
    :param corpus_words: corpus of processed words used to train the Word2Vec model.
    :param vector_size: the dimension of the vector of each word in document.
    :param workers:
    :param min_count:
    :param window:
    :return: a list (for the corpus) of the list (for every document) of vectors.
    """
    sub_start = time.time()

    model = Word2Vec(sentences=corpus_words, vector_size=vector_size, window=window, min_count=min_count, workers=workers)
    model.save(root_path + output_path + "word2vec.model")
    model.wv.save_word2vec_format(root_path + output_path + "word2vec_format.model")

    print("word2vec model build finished, use time " + str(time.time() - sub_start))
    return model


if __name__ == '__main__':
    start = time.time()
    train_word2vec_model(segment_and_clear(extract_label(read_excel("news.xlsx", "Sheet2"), "text1")))
